#!/usr/bin/python

import os
from openpyxl import load_workbook
from openpyxl import Workbook

def extract_data(filename, high):
	row_num = int(filename.split("_")[1].split(".")[0]) - 10000 
	new_row = []
	workbook = load_workbook(filename=filename)
	sheet = workbook.active
	if high:
		outsheet = wb["high"]
	else:
		outsheet = wb["low"]

	new_row.append((float(sheet["N35"].value)-float(sheet["B35"].value))-(float(sheet["T35"].value)-float(sheet["H35"].value)))
	new_row.append((float(sheet["O35"].value)-float(sheet["C35"].value)) -(float(sheet["U35"].value)-float(sheet["I35"].value)))
	new_row.append((float(sheet["P35"].value)-float(sheet["D35"].value))-(float(sheet["V35"].value)-float(sheet["J35"].value)))
	new_row.append((float(sheet["Q35"].value)-float(sheet["E35"].value))-(float(sheet["W35"].value)-float(sheet["K35"].value)))
	new_row.append((float(sheet["R35"].value)-float(sheet["F35"].value))-(float(sheet["X35"].value)-float(sheet["L35"].value)))
	new_row.append((float(sheet["S35"].value)-float(sheet["G35"].value))-(float(sheet["Y35"].value)-float(sheet["M35"].value)))
	new_row.append((float(sheet["AL35"].value)-float(sheet["Z35"].value))-(float(sheet["AR35"].value)-float(sheet["AF35"].value)))
	new_row.append((float(sheet["AM35"].value)-float(sheet["AA35"].value))-(float(sheet["AS35"].value)-float(sheet["AG35"].value)))
	new_row.append((float(sheet["AN35"].value)-float(sheet["AB35"].value))-(float(sheet["AT35"].value)-float(sheet["AH35"].value)))
	new_row.append((float(sheet["AO35"].value)-float(sheet["AC35"].value))-(float(sheet["AU35"].value)-float(sheet["AI35"].value)))
	new_row.append((float(sheet["AP35"].value)-float(sheet["AD35"].value))-(float(sheet["AV35"].value)-float(sheet["AJ35"].value)))
	new_row.append((float(sheet["AQ35"].value)-float(sheet["AE35"].value))-(float(sheet["AW35"].value)-float(sheet["AK35"].value)))	

	for col, val in enumerate(new_row, start=1):
		outsheet.cell(row=row_num, column=col).value = val

wb = Workbook()
wb.create_sheet(index = 0, title = "high")
wb.create_sheet(index = 1, title = "low")

for filename in os.listdir("."):
	if filename.startswith("high") and filename.endswith(".xlsx"):
		extract_data(filename, True)
	elif filename.startswith("low") and filename.endswith(".xlsx"):
		extract_data(filename, False)
	else:
		continue

wb.save("output.xlsx")
